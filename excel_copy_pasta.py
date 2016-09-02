# INSTRUCTIONS
# Be sure python is installed on your Windows computer (Python version 2.7 is ideal)
# MAKE SURE Excel file being read from is in the same folder as this file!
# Search for 'cmd' in the START menu to open the Command Prompt
#
# Run the following commands:
# cd ~
# cd path/to/where/you/saved/this/file
# pip install openpyxl
# python flatness_transpose.py
# Does not matter if file already exists or not.
#

from argparse import ArgumentParser
from openpyxl import load_workbook

# process the arguments from CLI
parser = ArgumentParser(description='Copy flatness data from one spreadsheet to the next. FILES MUST BE IN XLSX.')
parser.add_argument('-i', '--input', required=True, help='Filename of flatness data.')
parser.add_argument('-o', '--output', required=True, help='Filename to transpose to.')
parser.add_argument('-b', '--i_top', help='Column containing flatness bot data.', default='J')
parser.add_argument('-t', '--i_bot', help='Column containing flatness top data.', default='N')
parser.add_argument('-B', '--o_top', help='Column to output flatness bot data.', default='K')
parser.add_argument('-T', '--o_bot', help='Column to output flatness top data.', default='L')
parser.add_argument('-s', '--i_sheet', help='Sheet containing flatness data.', default='10053 - 10552')
parser.add_argument('-S', '--o_sheet', help='Sheet to output to.', default='Template')
args = parser.parse_args()
src = 'input_templates/%s' % (args.input)
dest = 'output_templates/%s' % (args.output)
# specify xlsx or xls file to read from
datawb = load_workbook(src, data_only=True)
flatSheet = datawb[args.i_sheet]  # specify sheet to read from

serialwb = load_workbook(dest)  # specify xlsx or xls file to read from
serialSheet = serialwb[args.o_sheet]  # specify sheet to read from

dataDict = {}
for row in range(2, flatSheet.max_row + 1):
    # make sure serial lookup key is UPPERCASE
    serial = flatSheet['A' + str(row)].value.upper()
    flattop = flatSheet[args.i_top + str(row)].value
    flatbot = flatSheet[args.i_bot + str(row)].value
    dataDict[serial] = [flattop, flatbot]
for row in range(2, serialSheet.max_row + 1):
    index = serialSheet['A' + str(row)].value
    if index in dataDict:
        # copy flat top to serialSheet
        serialSheet[args.o_top + str(row)] = dataDict[index][0]
        # copy flat bottom to serialSheet
        serialSheet[args.o_bot + str(row)] = dataDict[index][1]


serialwb.save(dest)
print 'Transpose Complete. Data entries transposed: %d' % (serialSheet.max_row)
